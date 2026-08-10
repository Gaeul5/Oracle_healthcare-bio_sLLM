"""
ADE 코퍼스 통합 변환기
=====================
CADECv2 / CADEC v1 / PsyTAR / PHEE -> 공통 스키마

설계 결정
---------
1. 불연속 엔티티: 최소 포함 구간으로 병합 (analysis_plan.md 결정 B)
   - 오프셋이 정렬돼 있지 않은 경우가 있으므로 반드시 sort 후 병합
   - 병합 후 동일 구간이 되는 엔티티는 중복 제거
2. 엔티티 매핑: MultiADE (Dai et al., 2024) Table 4의 행 #2를 따름
   n2c2 ADE = MADE ADE = PHEE Effect = PsyTAR ADR = CADEC ADE = CADECv2 ADE
   ※ PHEE의 'Adverse_event'는 트리거(예: "developed")이므로 사용하지 않음
3. 약물 그룹: MultiADE Table 2의 "26 drugs (4 groups)"에 맞춰
   NEXIUM과 PPIS를 ACID_RELATED 하나로 병합
4. 문서 단위: 게시글/초록 단위 (PsyTAR도 리뷰 단위)
   ※ MultiADE는 PsyTAR를 문장 단위(3,147)로 사용. 본 연구는 서사의
     인과 구조를 다루므로 리뷰 단위를 유지한다 (차이를 논문에 명시)

출력 스키마
-----------
doc_id, domain, corpus, drug_group, drug_name, text,
ade_spans, ade_offsets, n_discontig_merged, verbatim_ratio, orig_split
"""
import os
import re
import json
import collections
import pandas as pd

BASE = os.environ.get("ADE_DATA_ROOT", "/home/claude/work")
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")

# MultiADE Table 2 기준: 26 drugs, 4 groups
GROUP_MAP = {
    "NEXIUM": "ACID_RELATED",
    "PPIS": "ACID_RELATED",
    "CYMBALTA": "ANTIDEPRESSANT",
    "ANTIHYPERTENSIVES": "ANTIHYPERTENSIVE",
    "DICLOFENAC": "NSAID",
    # CADEC v1 전용
    "LIPITOR": "STATIN",
    # PsyTAR (전부 SSRI/SNRI)
    "LEXAPRO": "ANTIDEPRESSANT_PSY",
    "ZOLOFT": "ANTIDEPRESSANT_PSY",
    "CYMBALTA_PSY": "ANTIDEPRESSANT_PSY",
    "EFFEXORXR": "ANTIDEPRESSANT_PSY",
}

rows = []


def norm_text(t):
    return re.sub(r"\s+", " ", str(t).strip().lower())


def parse_brat_merged(ann_path, keep_types):
    """brat .ann 파싱 + 불연속 엔티티를 최소 포함 구간으로 병합(B안).

    Returns: (offsets, n_merged)  offsets = [(start, end), ...] 중복 제거됨
    """
    if not os.path.exists(ann_path):
        return [], 0
    seen, out, n_merged = set(), [], 0
    for line in open(ann_path, encoding="utf-8", errors="replace"):
        if not line.startswith("T"):
            continue
        parts = line.rstrip("\n").split("\t")
        if len(parts) < 3:
            continue
        bits = parts[1].split(" ", 1)
        if len(bits) < 2 or bits[0] not in keep_types:
            continue
        frags = []
        for seg in bits[1].split(";"):
            try:
                a, b = seg.split()
            except ValueError:
                continue
            frags.append((int(a), int(b)))
        if not frags:
            continue
        frags.sort()                       # ★ 정렬 없이 병합하면 빈 문자열이 나옴
        if len(frags) > 1:
            n_merged += 1
        span = (frags[0][0], frags[-1][1])
        if span in seen:                   # 병합으로 겹친 엔티티 제거
            continue
        seen.add(span)
        out.append(span)
    return sorted(out), n_merged


def _pack(doc_id, domain, corpus, group, drug, text, offsets,
          n_merged=0, spans=None, orig_split=""):
    if offsets:
        spans = [text[a:b] for a, b in offsets]
    spans = spans or []
    verbatim = sum(1 for s in spans if s and s.lower() in text.lower())
    rows.append(dict(
        doc_id=doc_id, domain=domain, corpus=corpus,
        drug_group=GROUP_MAP.get(group, group), drug_name=drug,
        text=text, ade_spans=spans, ade_offsets=offsets,
        n_discontig_merged=n_merged,
        verbatim_ratio=round(verbatim / len(spans), 3) if spans else None,
        orig_split=orig_split,
    ))


# ------------------------------------------------------------------ CADECv2
def load_cadecv2():
    root = f"{BASE}/cadecv2_new/data/extracted/data"
    split_of = {}
    for sp in ("train", "dev", "test"):
        p = f"{root}/split/{sp}.id"
        if os.path.exists(p):
            for line in open(p):
                split_of[line.strip()] = sp
    n = 0
    for fn in sorted(os.listdir(f"{root}/txt")):
        if not fn.endswith(".txt"):
            continue
        stem = fn[:-4]
        base = stem.rsplit(".", 1)[0]
        cls, drug = base.split("-", 1) if "-" in base else (base, base)
        text = open(f"{root}/txt/{fn}", encoding="utf-8", errors="replace").read()
        offs, nm = parse_brat_merged(f"{root}/ann/{stem}.ann", {"ADR"})
        _pack(f"cadecv2:{stem}", "forum", "cadecv2", cls, drug, text,
              offs, nm, orig_split=split_of.get(stem, ""))
        n += 1
    return n


# ----------------------------------------------------------------- CADEC v1
def load_cadec_v1():
    root = f"{BASE}/cadec/v2/cadec"
    n = 0
    for fn in sorted(os.listdir(f"{root}/text")):
        if not fn.endswith(".txt"):
            continue
        stem = fn[:-4]
        drug = stem.rsplit(".", 1)[0]
        # v1은 유효성분이 2종뿐: LIPITOR(atorvastatin) 외 전부 diclofenac 계열
        cls = "LIPITOR" if drug.upper().startswith("LIPITOR") else "DICLOFENAC"
        text = open(f"{root}/text/{fn}", encoding="utf-8", errors="replace").read()
        offs, nm = parse_brat_merged(f"{root}/original/{stem}.ann", {"ADR"})
        _pack(f"cadecv1:{stem}", "forum", "cadec_v1", cls, drug, text, offs, nm)
        n += 1
    return n


# ------------------------------------------------------------------- PsyTAR
def load_psytar():
    import openpyxl
    wb = openpyxl.load_workbook(
        f"{BASE}/psytar/psytar/PsyTAR_dataset.xlsx", read_only=True)

    # ★ 주석은 (side-effect 필드 + comment 필드)를 문장 분할한 텍스트에 정렬된다.
    #   Sample.comment만 쓰면 부작용 표현의 80%가 원문에서 안 잡힌다.
    #   따라서 Sentence_Labeling 시트의 문장을 순서대로 이어붙여 문서를 복원한다.
    wsS = wb["Sentence_Labeling"]
    hS = [c.value for c in next(wsS.iter_rows(min_row=1, max_row=1))]
    dS, iS, sS = hS.index("drug_id"), hS.index("sentence_index"), hS.index("sentences")
    buf = collections.defaultdict(list)
    for r in wsS.iter_rows(min_row=2, values_only=True):
        if r[dS] and r[sS] and str(r[sS]).strip():
            try:
                idx = int(r[iS])
            except (TypeError, ValueError):
                idx = len(buf[str(r[dS])])
            buf[str(r[dS])].append((idx, str(r[sS]).strip()))
    text_of = {
        did: " ".join(s for _, s in sorted(sents))
        for did, sents in buf.items()
    }

    ws2 = wb["ADR_Identified"]
    h2 = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
    d2 = h2.index("drug_id")
    adr_cols = [i for i, c in enumerate(h2) if c and str(c).startswith("ADR")]
    adrs = collections.defaultdict(list)
    for r in ws2.iter_rows(min_row=2, values_only=True):
        if not r[d2]:
            continue
        for i in adr_cols:
            if i < len(r) and r[i] and str(r[i]).strip():
                adrs[str(r[d2])].append(str(r[i]).strip())

    n = 0
    for did, text in text_of.items():
        drug = did.split(".")[0].upper()
        key = "CYMBALTA_PSY" if drug == "CYMBALTA" else drug
        spans = list(dict.fromkeys(adrs.get(did, [])))
        # PsyTAR은 오프셋 미제공 -> 원문에서 역탐색 (실패 시 빈 오프셋)
        offs = []
        low = text.lower()
        for s in spans:
            i = low.find(s.lower())
            if i >= 0:
                offs.append((i, i + len(s)))
        _pack(f"psytar:{did}", "forum", "psytar", key, drug, text,
              offsets=[], spans=spans)
        rows[-1]["ade_offsets"] = sorted(offs)
        n += 1
    return n


# --------------------------------------------------------------------- PHEE
def load_phee():
    root = f"{BASE}/PHEE-master/PHEE-master/data/json"
    n = 0
    for sp in ("train", "dev", "test"):
        for line in open(f"{root}/{sp}.json", encoding="utf-8"):
            if not line.strip():
                continue
            r = json.loads(line)
            eff, drugs = [], []
            for a in r.get("annotations", []):
                for e in a.get("events", []):
                    if e.get("event_type") != "Adverse_event":
                        continue
                    # MultiADE Table 4 #2: PHEE의 'Effect'를 ADE로 매핑
                    # ('Adverse_event' 트리거는 별도 카테고리 #15이므로 제외)
                    for grp in (e.get("Effect") or {}).get("text", []):
                        eff += [str(t).strip() for t in grp if str(t).strip()]
                    dr = (e.get("Treatment") or {}).get("Drug") or {}
                    for grp in dr.get("text", []):
                        drugs += [str(t).strip().lower() for t in grp if str(t).strip()]
            text = r["context"]
            spans = list(dict.fromkeys(eff))
            offs, low = [], text.lower()
            for s in spans:
                i = low.find(s.lower())
                if i >= 0:
                    offs.append((i, i + len(s)))
            _pack(f"phee:{r['id']}", "literature", "phee", "LITERATURE",
                  "|".join(dict.fromkeys(drugs)), text,
                  offsets=[], spans=spans, orig_split=sp)
            rows[-1]["ade_offsets"] = sorted(offs)
            n += 1
    return n


def main():
    counts = {
        "cadecv2": load_cadecv2(),
        "cadec_v1": load_cadec_v1(),
        "psytar": load_psytar(),
        "phee": load_phee(),
    }
    for k, v in counts.items():
        print(f"{k:10s}: {v}")

    df = pd.DataFrame(rows)

    # 전역 중복 제거 (본문 기준). 우선순위: 주석 품질이 높은 순
    prio = {"cadecv2": 0, "psytar": 1, "cadec_v1": 2, "phee": 3}
    df["_n"] = df.text.map(norm_text)
    df["_p"] = df.corpus.map(prio)
    before = len(df)
    df = df.sort_values("_p", kind="stable").drop_duplicates("_n", keep="first").sort_index()
    print(f"\n중복 제거: {before} -> {len(df)}")
    df = df.drop(columns=["_n", "_p"])
    df["n_ade"] = df.ade_spans.map(len)

    os.makedirs(OUT, exist_ok=True)
    path = os.path.join(OUT, "unified.jsonl")
    df.to_json(path, orient="records", lines=True, force_ascii=False)
    print(f"저장: {path}  ({len(df)}건)")
    return df


if __name__ == "__main__":
    main()
